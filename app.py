from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import os, io, json, base64
from datetime import datetime
import calendar
from report_generator import generate_report
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['REPORT_FOLDER'] = 'reports'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
TRAINING_DATA_PATH = 'data/final_dataset_.xlsx'
LOGO_PATH = 'static/img/logo.png'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['REPORT_FOLDER'], exist_ok=True)

# ─── Data Engine ─────────────────────────────────────────────────────────────
class DataEngine:
    def __init__(self, df: pd.DataFrame):
        self.df = self._prepare(df)

    def _prepare(self, df: pd.DataFrame) -> pd.DataFrame:
        df.columns = df.columns.astype(str).str.strip()
        date_cols = ['Received_Date', 'Forward_Date', 'Return_Date', 'NSN_Allotment_Date']
        for c in date_cols:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], dayfirst=True, errors='coerce')
        if 'MRC' in df.columns:
            df['MRC'] = pd.to_numeric(df['MRC'], errors='coerce')
        if 'DPSU' in df.columns:
            df['DPSU'] = df['DPSU'].astype(str).str.strip()
        if 'NCB' in df.columns:
            df['NCB'] = df['NCB'].astype(str).str.strip()
        return df

    def filter(self, start, end, col='Received_Date'):
        m = (self.df[col] >= pd.Timestamp(start)) & (self.df[col] <= pd.Timestamp(end))
        return self.df[m]

    @staticmethod
    def monthly_ranges(year, month):
        first = datetime(year, month, 1)
        last  = datetime(year, month, calendar.monthrange(year, month)[1])
        return first, last

    def summary(self, filtered_df):
        total    = len(filtered_df)
        fwd      = filtered_df['Forward_Date'].notna().sum()       if 'Forward_Date'       in filtered_df.columns else 0
        nsn      = filtered_df['NSN_Allotment_Date'].notna().sum() if 'NSN_Allotment_Date' in filtered_df.columns else 0
        returned = filtered_df['Return_Date'].notna().sum()        if 'Return_Date'        in filtered_df.columns else 0
        if 'Forward_Date' in filtered_df.columns and 'NSN' in filtered_df.columns and 'NSN_Allotment_Date' in filtered_df.columns:
            pending = filtered_df[(filtered_df['Forward_Date'].notna()) &
                                  (filtered_df['NSN'].isna()) &
                                  (filtered_df['NSN_Allotment_Date'].isna())].shape[0]
        else:
            pending = 0
        by_dpsu  = filtered_df.groupby('DPSU').size().to_dict() if 'DPSU' in filtered_df.columns else {}
        by_ncb   = filtered_df.groupby('NCB').size().to_dict()  if 'NCB'  in filtered_df.columns else {}
        by_equip = filtered_df.groupby('Equipment_Name').size().nlargest(10).to_dict() if 'Equipment_Name' in filtered_df.columns else {}

        if 'DPSU' in filtered_df.columns:
            fwd_by_dpsu = filtered_df[filtered_df['Forward_Date'].notna()].groupby('DPSU').size().to_dict() if 'Forward_Date' in filtered_df.columns else {}
            ret_by_dpsu = filtered_df[filtered_df['Return_Date'].notna()].groupby('DPSU').size().to_dict() if 'Return_Date' in filtered_df.columns else {}
            # ── NEW: NSN allotted per DPSU ──
            nsn_by_dpsu = filtered_df[filtered_df['NSN_Allotment_Date'].notna()].groupby('DPSU').size().to_dict() if 'NSN_Allotment_Date' in filtered_df.columns else {}
            if 'NSN' in filtered_df.columns and 'NSN_Allotment_Date' in filtered_df.columns:
                pend_by_dpsu = filtered_df[(filtered_df['Forward_Date'].notna()) &
                                           (filtered_df['NSN'].isna()) &
                                           (filtered_df['NSN_Allotment_Date'].isna())].groupby('DPSU').size().to_dict()
            else:
                pend_by_dpsu = {}
        else:
            fwd_by_dpsu = ret_by_dpsu = pend_by_dpsu = nsn_by_dpsu = {}

        if 'MRC' in filtered_df.columns and total:
            mrc_vals = filtered_df['MRC'].dropna()
            avg_mrc = round(mrc_vals.mean(), 2) if len(mrc_vals) else 0
        else:
            avg_mrc = 0
        if 'Received_Date' in filtered_df.columns and 'Forward_Date' in filtered_df.columns:
            proc = (filtered_df['Forward_Date'] - filtered_df['Received_Date']).dt.days.dropna()
            avg_proc = round(proc.mean(), 1) if len(proc) else 0
        else:
            avg_proc = 0
        return {
            'total': int(total), 'forwarded': int(fwd), 'nsn_allotted': int(nsn),
            'returned': int(returned), 'pending': int(pending),
            'by_dpsu':  {k: int(v) for k, v in by_dpsu.items()},
            'by_ncb':   {k: int(v) for k, v in by_ncb.items()},
            'by_equipment': {k: int(v) for k, v in by_equip.items()},
            'fwd_by_dpsu':  {k: int(v) for k, v in fwd_by_dpsu.items()},
            'ret_by_dpsu':  {k: int(v) for k, v in ret_by_dpsu.items()},
            'pend_by_dpsu': {k: int(v) for k, v in pend_by_dpsu.items()},
            'nsn_by_dpsu':  {k: int(v) for k, v in nsn_by_dpsu.items()},  # NEW
            'avg_mrc': float(avg_mrc), 'avg_processing_days': float(avg_proc),
        }

    def group_for_report(self, filtered_df):
        rows = []
        if filtered_df.empty:
            return rows
        grp = filtered_df.groupby(['DPSU', 'Equipment_Name'])
        for (dpsu, equip), sub in grp:
            rows.append({
                'dpsu': dpsu, 'equipment': equip,
                'total_items': len(sub),
                'forwarded': int(sub['Forward_Date'].notna().sum()) if 'Forward_Date' in sub.columns else 0,
                'nsn_allotted': int(sub['NSN_Allotment_Date'].notna().sum()) if 'NSN_Allotment_Date' in sub.columns else 0,
                'returned': int(sub['Return_Date'].notna().sum()) if 'Return_Date' in sub.columns else 0,
                'pending': len(sub) - int(sub['Return_Date'].notna().sum() if 'Return_Date' in sub.columns else 0),
                'avg_mrc': round(sub['MRC'].mean(), 1) if 'MRC' in sub.columns else 0,
            })
        return rows


# ─── Colour Palette ──────────────────────────────────────────────────────────
C = {
    'navy':      '0D2240',
    'navy_mid':  '1B3A6B',
    'navy_lt':   '2E6DA4',
    'gold':      'FFD700',
    'gold_dk':   'B8860B',
    'white':     'FFFFFF',
    'off_white': 'F5F8FF',
    'gray_lt':   'E8EEF6',
    'gray':      'C5D0E0',
    'red':       'C0392B',
    'red_lt':    'E74C3C',
    'green':     '1A7A4A',
    'green_lt':  '27AE60',
    'alt1':      'EBF4FF',
    'alt2':      'FFFFFF',
    'total_bg':  '0A1929',
    'accent':    'FFD700',
    'crimson':   '8B0000',
}

def _side(color='AAAAAA', style='thin'):
    return Side(style=style, color=color)

def thin_border(color='BBBBBB'):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = _side('1B3A6B', 'medium')
    return Border(left=s, right=s, top=s, bottom=s)

def gold_border():
    s = _side('B8860B', 'medium')
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):
    return PatternFill('solid', fgColor=color)

def style_cell(c, value=None, bold=False, fg='000000', bg=None, size=9,
               align='center', valign='center', wrap=False, border=True,
               italic=False, num_fmt=None, underline=False):
    if value is not None:
        c.value = value
    c.font = Font(bold=bold, italic=italic, color=fg, size=size, name='Calibri',
                  underline='single' if underline else None)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


# ─── Excel Export from report_data ────────────────────────────────────────────
def build_excel_from_report_data(report_data, heading, filename_base):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Codification Report'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    col_widths = [6, 22, 38, 18, 20, 18, 14, 14, 14, 18, 16, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.merge_cells('A1:M1')
    c = ws['A1']
    c.value = heading
    c.font  = Font(bold=True, color=C['white'], size=12, name='Calibri')
    c.fill  = fill(C['navy_mid'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28

    group_hdrs = [
        ('A2:A3',  'S.No'),
        ('B2:B3',  'AHSP / DPSU'),
        ('C2:C3',  'Equipment Name'),
        ('D2:D3',  'Codification\nTarget (25-27)'),
        ('E2:E3',  'As per DPSUs/\nAHSP MRLs'),
        ('F2:I2',  'PROGRESS'),
        ('J2:J3',  'Updation\nTarget (25-27)'),
        ('K2:K3',  'Updation Done\nby AHSPs / DPSUs'),
        ('L2:L3',  '% Updated'),
        ('M2:M3',  'Remarks'),
    ]
    for merge_r, label in group_hdrs:
        ws.merge_cells(merge_r)
        start = merge_r.split(':')[0]
        c = ws[start]
        c.value = label
        c.font  = Font(bold=True, color=C['gold'], size=8, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border()

    sub_hdrs = [
        ('F3', 'Total Items Codified\nby AHSPs / DPSUs'),
        ('G3', 'Fwd to\nDCA'),
        ('H3', 'NSN\nAllotted'),
        ('I3', 'Returned to\nAHSPs / DPSUs'),
    ]
    for cell_ref, label in sub_hdrs:
        c = ws[cell_ref]
        c.value = label
        c.font  = Font(bold=True, color=C['white'], size=7, name='Calibri')
        c.fill  = fill(C['navy'])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border()

    current_row = 4
    sno = 1
    total_codified = 0
    total_fwd = 0
    total_nsn = 0
    total_returned = 0

    for dpsu, items in report_data.items():
        for item_idx, item in enumerate(items):
            bg = C['alt1'] if current_row % 2 == 0 else C['alt2']
            ws.row_dimensions[current_row].height = 18

            c = ws.cell(row=current_row, column=1)
            style_cell(c, sno, bold=True, fg=C['navy_mid'], bg=bg, size=9)

            c = ws.cell(row=current_row, column=2)
            style_cell(c, dpsu if item_idx == 0 else '', bold=(item_idx == 0),
                       fg=C['navy_mid'], bg=bg, align='left', wrap=True, size=8)

            c = ws.cell(row=current_row, column=3)
            style_cell(c, item['Equipment'], bg=bg, align='left', wrap=True, size=8, fg='1A1A2E')

            c = ws.cell(row=current_row, column=4)
            style_cell(c, '', bg=bg, fg='888888', size=8)

            c = ws.cell(row=current_row, column=5)
            style_cell(c, '', bg=bg, fg='888888', size=8)

            c = ws.cell(row=current_row, column=6)
            val = item['Total_Codified']
            style_cell(c, val, bg=bg, fg=C['green'] if val > 0 else '999999', bold=(val > 0), size=9)

            c = ws.cell(row=current_row, column=7)
            val = item['Fwd_DCA']
            style_cell(c, val, bg=bg, fg=C['green'] if val > 0 else '999999', bold=(val > 0), size=9)

            c = ws.cell(row=current_row, column=8)
            val = item['NSN']
            style_cell(c, val, bg=bg, fg=C['green'] if val > 0 else '999999', bold=(val > 0), size=9)

            c = ws.cell(row=current_row, column=9)
            val = item['Returned']
            style_cell(c, val, bg=bg, fg=C['green'] if val > 0 else '999999', bold=(val > 0), size=9)

            for col_off in [10, 11, 12, 13]:
                c = ws.cell(row=current_row, column=col_off)
                style_cell(c, '', bg=bg, fg='888888', size=8)

            total_codified += item['Total_Codified']
            total_fwd      += item['Fwd_DCA']
            total_nsn      += item['NSN']
            total_returned += item['Returned']
            sno += 1
            current_row += 1

    ws.row_dimensions[current_row].height = 24

    c = ws.cell(row=current_row, column=1)
    c.value = sno
    c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
    c.fill  = fill(C['total_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = gold_border()

    ws.merge_cells(f'B{current_row}:E{current_row}')
    c = ws[f'B{current_row}']
    c.value = 'TOTAL'
    c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
    c.fill  = fill(C['total_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = gold_border()

    total_vals = [total_codified, total_fwd, total_nsn, total_returned]
    for col_off, val in enumerate(total_vals, 6):
        c = ws.cell(row=current_row, column=col_off)
        c.value = val
        c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    for col_off in [10, 11, 12, 13]:
        c = ws.cell(row=current_row, column=col_off)
        c.value = ''
        c.font  = Font(bold=True, color=C['gray'], size=10, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    sig_row = current_row + 3
    sig_data = [
        ('A', 'D', 'Prepared By'),
        ('E', 'I', 'Checked By / DS Member'),
        ('J', 'M', 'Approved By / Dir Std'),
    ]
    for sc, ec, label in sig_data:
        ws.merge_cells(f'{sc}{sig_row}:{ec}{sig_row}')
        c = ws[f'{sc}{sig_row}']
        style_cell(c, '', bg=C['gray_lt'], border=True, fg='666666', size=8)
        ws.row_dimensions[sig_row].height = 14

        ws.merge_cells(f'{sc}{sig_row+1}:{ec}{sig_row+1}')
        ws.row_dimensions[sig_row+1].height = 35
        c = ws[f'{sc}{sig_row+1}']
        c.fill = fill(C['off_white'])
        c.border = thin_border()

        ws.merge_cells(f'{sc}{sig_row+2}:{ec}{sig_row+2}')
        ws.row_dimensions[sig_row+2].height = 18
        c = ws[f'{sc}{sig_row+2}']
        style_cell(c, '_' * 30, fg='333333', bg=C['off_white'], border=True, size=9)

        ws.merge_cells(f'{sc}{sig_row+3}:{ec}{sig_row+3}')
        ws.row_dimensions[sig_row+3].height = 14
        c = ws[f'{sc}{sig_row+3}']
        style_cell(c, label, bold=True, fg=C['navy_mid'], bg=C['gray_lt'],
                   border=True, size=9, align='center')

    out_path = os.path.join('reports', f'{filename_base}.xlsx')
    wb.save(out_path)
    return out_path


# ─── Main Report Builder ─────────────────────────────────────────────────────
def _build_report(stats, rows, title, subtitle, filename_base, report_type, period_label):
    wb = Workbook()

    ws = wb.active
    ws.title = 'Codification Summary'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    col_widths = [5, 22, 42, 14, 18, 16, 13, 13, 13, 14, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width  = 160
            img.height = 110
            img.anchor = 'A1'
            ws.add_image(img)
        except Exception:
            pass

    for r, h in [(1,28),(2,28),(3,28),(4,28),(5,36),(6,28),(7,40),(8,46)]:
        ws.row_dimensions[r].height = h

    ws.merge_cells('C1:M1')
    c = ws['C1']
    c.value = 'भारत सरकार / GOVERNMENT OF INDIA'
    c.font  = Font(bold=True, color=C['gold'], size=13, name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C2:M2')
    c = ws['C2']
    c.value = 'रक्षा मंत्रालय / MINISTRY OF DEFENCE'
    c.font  = Font(bold=True, color=C['white'], size=11, name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C3:M3')
    c = ws['C3']
    c.value = 'मानकीकरण निदेशालय, सदस्य ए सी/135'
    c.font  = Font(bold=True, color=C['gold'], size=11, name='Calibri', italic=True)
    c.fill  = fill(C['navy_mid'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C4:M4')
    c = ws['C4']
    c.value = 'DIRECTORATE OF STANDARDISATION, MEMBER AC/135'
    c.font  = Font(bold=True, color=C['off_white'], size=10, name='Calibri')
    c.fill  = fill(C['navy_mid'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    for r in range(1, 5):
        for col in ['A', 'B']:
            ws[f'{col}{r}'].fill = fill(C['navy'])

    ws.merge_cells('A5:M5')
    c = ws['A5']
    c.value = '★  CODIFICATION INTELLIGENCE SYSTEM  ★'
    c.font  = Font(bold=True, color=C['navy'], size=10, name='Calibri', italic=True)
    c.fill  = fill(C['gold'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A6:M6')
    c = ws['A6']
    c.value = title
    c.font  = Font(bold=True, color=C['white'], size=14, name='Calibri')
    c.fill  = fill(C['navy_mid'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A7:M7')
    c = ws['A7']
    c.value = subtitle
    c.font  = Font(bold=True, color=C['gold'], size=11, name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A8:D8')
    c = ws['A8']
    c.value = f'Report Type: {report_type}   |   Period: {period_label}'
    c.font  = Font(bold=True, size=8, color=C['navy_mid'], name='Calibri')
    c.fill  = fill(C['gray_lt'])
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border('CCCCCC')

    ws.merge_cells('E8:I8')
    c = ws['E8']
    c.value = f'Generated: {datetime.now().strftime("%d %B %Y, %H:%M hrs")}'
    c.font  = Font(size=8, color='555555', name='Calibri')
    c.fill  = fill(C['gray_lt'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border('CCCCCC')

    ws.merge_cells('J8:M8')
    c = ws['J8']
    c.value = f'Avg Processing: {stats["avg_processing_days"]} days'
    c.font  = Font(bold=True, size=8, color=C['navy_mid'], name='Calibri')
    c.fill  = fill(C['gray_lt'])
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = thin_border('CCCCCC')

    ws.row_dimensions[9].height = 14
    ws.row_dimensions[10].height = 38

    kpi_label_ranges = ['A9:C9','D9:F9','G9:I9','J9:K9','L9:M9']
    kpi_labels = ['TOTAL ITEMS','FWD TO DCA','NSN ALLOTTED','RETURNED','PENDING']
    kpi_bgs    = [C['navy_mid'], C['navy_lt'], C['green'], '1D6A40', C['red']]
    for merge_range, label, bg in zip(kpi_label_ranges, kpi_labels, kpi_bgs):
        ws.merge_cells(merge_range)
        start = merge_range.split(':')[0]
        c = ws[start]
        c.value = label
        c.font  = Font(bold=True, color=C['gold'], size=7, name='Calibri')
        c.fill  = fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    kpi_vals = [stats['total'], stats['forwarded'], stats['nsn_allotted'], stats['returned'], stats['pending']]
    kpi_val_ranges = ['A10:C10','D10:F10','G10:I10','J10:K10','L10:M10']
    for merge_range, val, bg in zip(kpi_val_ranges, kpi_vals, kpi_bgs):
        ws.merge_cells(merge_range)
        start = merge_range.split(':')[0]
        c = ws[start]
        c.value = val
        c.font  = Font(bold=True, color=C['white'], size=20, name='Calibri')
        c.fill  = fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    ws.row_dimensions[11].height = 12
    ws.row_dimensions[12].height = 14
    ws.row_dimensions[13].height = 42

    ws.merge_cells('A11:M11')
    ws['A11'].fill = fill(C['gold'])

    group_hdrs = [
        ('A12:A13', 'S.No'),
        ('B12:B13', 'AsHSP / DPSU'),
        ('C12:C13', 'NAME OF EQUIPMENT'),
        ('D12:E12', 'TARGETS (25-27)'),
        ('F12:I12', 'CODIFICATION PROGRESS'),
        ('J12:K12', 'UPDATION'),
        ('L12:L13', 'REMARKS'),
    ]
    for merge_r, label in group_hdrs:
        ws.merge_cells(merge_r)
        start = merge_r.split(':')[0]
        c = ws[start]
        c.value = label
        c.font  = Font(bold=True, color=C['gold'], size=8, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border()

    sub_hdrs = [
        ('D13', 'Codification\nTarget'),
        ('E13', 'As per DPSUs/\nAHSP MRLs'),
        ('F13', 'Total Items Codified\nby AHSPs / DPSUs'),
        ('G13', 'Fwd to\nDCA'),
        ('H13', 'NSN\nAllotted'),
        ('I13', 'Returned to\nAHSPs / DPSUs'),
        ('J13', 'Updation\nTarget'),
        ('K13', 'Updation Done\nby AHSPs / DPSUs'),
    ]
    for cell_ref, label in sub_hdrs:
        c = ws[cell_ref]
        c.value = label
        c.font  = Font(bold=True, color=C['white'], size=7, name='Calibri')
        c.fill  = fill(C['navy'])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border()

    current_row = 14
    dpsu_groups = {}
    for r in rows:
        dpsu_groups.setdefault(r['dpsu'], []).append(r)

    sno = 1
    for dpsu, items in dpsu_groups.items():
        for item_idx, item in enumerate(items):
            bg = C['alt1'] if current_row % 2 == 0 else C['alt2']
            ws.row_dimensions[current_row].height = 18

            c = ws.cell(row=current_row, column=1)
            style_cell(c, sno if item_idx == 0 else '', bold=True, fg=C['navy_mid'], bg=bg, size=9)

            c = ws.cell(row=current_row, column=2)
            style_cell(c, dpsu if item_idx == 0 else '', bold=(item_idx==0),
                       fg=C['navy_mid'] if item_idx==0 else '555555', bg=bg, align='left', wrap=True, size=8)

            c = ws.cell(row=current_row, column=3)
            style_cell(c, item['equipment'], bg=bg, align='left', wrap=True, size=8, fg='1A1A2E')

            for col in [4, 5]:
                c = ws.cell(row=current_row, column=col)
                style_cell(c, '', bg=bg, fg='888888', size=8)

            prog_vals = [item['total_items'], item['forwarded'], item['nsn_allotted'], item['returned']]
            for col_off, val in enumerate(prog_vals, 6):
                c = ws.cell(row=current_row, column=col_off)
                vfg = C['green'] if val > 0 else '999999'
                style_cell(c, val, bg=bg, fg=vfg, bold=(val > 0), size=9)

            for col in [10, 11, 12]:
                c = ws.cell(row=current_row, column=col)
                style_cell(c, '', bg=bg, fg='888888', size=8)

            current_row += 1
        sno += 1

    ws.row_dimensions[current_row].height = 24
    ws.merge_cells(f'A{current_row}:C{current_row}')
    c = ws[f'A{current_row}']
    c.value = 'G R A N D   T O T A L'
    c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
    c.fill  = fill(C['total_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = gold_border()

    total_data = ['', '', stats['total'], stats['forwarded'], stats['nsn_allotted'],
                  stats['returned'], '', '', '']
    for col_off, val in enumerate(total_data, 4):
        c = ws.cell(row=current_row, column=col_off)
        is_val = isinstance(val, (int, float)) and col_off not in [4, 5, 10, 11]
        c.value = val
        c.font  = Font(bold=True, color=C['gold'] if is_val else C['gray'], size=10, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    sig_row = current_row + 3
    sig_data = [
        ('A', 'D', 'Prepared By'),
        ('E', 'I', 'Checked By / DS Member'),
        ('J', 'M', 'Approved By / Dir Std'),
    ]
    for sc, ec, label in sig_data:
        ws.merge_cells(f'{sc}{sig_row}:{ec}{sig_row}')
        c = ws[f'{sc}{sig_row}']
        style_cell(c, '', bg=C['gray_lt'], border=True, fg='666666', size=8)

        ws.merge_cells(f'{sc}{sig_row+1}:{ec}{sig_row+1}')
        ws.row_dimensions[sig_row+1].height = 35
        c = ws[f'{sc}{sig_row+1}']
        c.fill = fill(C['off_white'])
        c.border = thin_border()

        ws.merge_cells(f'{sc}{sig_row+2}:{ec}{sig_row+2}')
        ws.row_dimensions[sig_row+2].height = 18
        c = ws[f'{sc}{sig_row+2}']
        style_cell(c, '_' * 30, fg='333333', bg=C['off_white'], border=True, size=9)

        ws.merge_cells(f'{sc}{sig_row+3}:{ec}{sig_row+3}')
        c = ws[f'{sc}{sig_row+3}']
        style_cell(c, label, bold=True, fg=C['navy_mid'], bg=C['gray_lt'],
                   border=True, size=9, align='center')

    foot_row = sig_row + 5
    ws.merge_cells(f'A{foot_row}:M{foot_row}')
    c = ws[f'A{foot_row}']
    c.value = (f'DEFENCE STANDARDISATION, BENGALURU  |  Member AC/135  |  '
               f'Directorate of Standardisation  |  Ministry of Defence, Govt. of India  |  '
               f'Developed by: Cobra Tech, ABVGIET Shimla')
    c.font  = Font(italic=True, size=7, color='888888', name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    # SHEET 2: DPSU BREAKDOWN
    ws2 = wb.create_sheet('DPSU Analysis')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_view.zoomScale = 90
    col_w2 = [6, 26, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    if os.path.exists(LOGO_PATH):
        try:
            img2 = XLImage(LOGO_PATH)
            img2.width  = 120
            img2.height = 83
            img2.anchor = 'A1'
            ws2.add_image(img2)
        except Exception:
            pass

    for r in range(1, 4):
        ws2.row_dimensions[r].height = 26
        for col in ['A', 'B']:
            ws2[f'{col}{r}'].fill = fill(C['navy'])

    ws2.merge_cells('C1:G1')
    c = ws2['C1']
    c.value = 'DIRECTORATE OF STANDARDISATION — BENGALURU'
    c.font  = Font(bold=True, color=C['gold'], size=12, name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('C2:G2')
    c = ws2['C2']
    c.value = title
    c.font  = Font(bold=True, color=C['white'], size=10, name='Calibri')
    c.fill  = fill(C['navy_mid'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('C3:G3')
    c = ws2['C3']
    c.value = 'DPSU / AsHSP — WISE CODIFICATION ANALYSIS'
    c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
    c.fill  = fill(C['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('A4:G4')
    ws2['A4'].fill = fill(C['gold'])
    ws2.row_dimensions[4].height = 8
    ws2.row_dimensions[5].height = 32

    h2_labels = ['S.No', 'DPSU / AsHSP', 'Total Items', 'NSN Allotted', 'Returned', 'Pending']
    for col_idx, label in enumerate(h2_labels, 1):
        c = ws2.cell(row=5, column=col_idx)
        c.value = label
        c.font  = Font(bold=True, color=C['gold'], size=9, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border()

    row2 = 6
    for s_idx, (dpsu, cnt) in enumerate(stats['by_dpsu'].items(), 1):
        ws2.row_dimensions[row2].height = 20
        bg = C['alt1'] if row2 % 2 == 0 else C['alt2']
        dpsu_rows = [x for x in rows if x['dpsu'] == dpsu]
        tot_nsn   = sum(x['nsn_allotted'] for x in dpsu_rows)
        tot_ret   = sum(x['returned']     for x in dpsu_rows)
        tot_pend  = cnt - tot_ret
        row_vals = [s_idx, dpsu, cnt, tot_nsn, tot_ret, tot_pend]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws2.cell(row=row2, column=col_idx)
            c.value = val
            c.font  = Font(size=9, bold=(col_idx <= 2), name='Calibri',
                           color=(C['navy_mid'] if col_idx == 2 else '1A1A1A'))
            c.fill  = fill(bg)
            c.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left', vertical='center')
            c.border = thin_border()
        row2 += 1

    ws2.row_dimensions[row2].height = 22
    ws2.merge_cells(f'A{row2}:B{row2}')
    c = ws2[f'A{row2}']
    style_cell(c, 'TOTAL', bold=True, fg=C['gold'], bg=C['total_bg'], size=10, border=False)
    c.border = gold_border()
    tot2_vals = [stats['total'], stats['nsn_allotted'], stats['returned'], stats['pending']]
    for col_idx, val in enumerate(tot2_vals, 3):
        c = ws2.cell(row=row2, column=col_idx)
        c.value = val
        c.font  = Font(bold=True, color=C['gold'], size=10, name='Calibri')
        c.fill  = fill(C['total_bg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = gold_border()

    out_path = os.path.join('reports', f'{filename_base}.xlsx')
    wb.save(out_path)
    return out_path


# ─── Routes ──────────────────────────────────────────────────────────────────
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only Excel files supported'}), 400
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    f.save(path)
    try:
        df = pd.read_excel(path)
        cols = list(df.columns)
        preview = df.head(5).fillna('').to_dict(orient='records')
        return jsonify({'success': True, 'columns': cols, 'rows': len(df),
                        'preview': preview, 'filename': f.filename})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/analyze', methods=['POST'])
def analyze():
    data = request.json
    filename = data.get('filename')
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename) if filename else None
    if not path or not os.path.exists(path):
        path = TRAINING_DATA_PATH
    if not os.path.exists(path):
        return jsonify({'error': 'No dataset available'}), 404
    df = pd.read_excel(path)
    engine = DataEngine(df)
    stats = engine.summary(df)
    if 'Received_Date' in df.columns:
        dates = pd.to_datetime(df['Received_Date'], dayfirst=True, errors='coerce').dropna()
        stats['date_min'] = dates.min().strftime('%d-%b-%Y') if len(dates) else 'N/A'
        stats['date_max'] = dates.max().strftime('%d-%b-%Y') if len(dates) else 'N/A'
    else:
        stats['date_min'] = stats['date_max'] = 'N/A'
    return jsonify(stats)

@app.route('/analyze_multi', methods=['POST'])
def analyze_multi():
    """Analyze multiple files and return stats for each."""
    data = request.json
    filenames = data.get('filenames', [])
    results = []
    for filename in filenames:
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename) if filename else None
        if not path or not os.path.exists(path):
            path = TRAINING_DATA_PATH
        if not os.path.exists(path):
            results.append({'filename': filename, 'error': 'File not found'})
            continue
        try:
            df = pd.read_excel(path)
            engine = DataEngine(df)
            stats = engine.summary(df)
            if 'Received_Date' in df.columns:
                dates = pd.to_datetime(df['Received_Date'], dayfirst=True, errors='coerce').dropna()
                stats['date_min'] = dates.min().strftime('%d-%b-%Y') if len(dates) else 'N/A'
                stats['date_max'] = dates.max().strftime('%d-%b-%Y') if len(dates) else 'N/A'
            else:
                stats['date_min'] = stats['date_max'] = 'N/A'
            stats['filename'] = filename
            results.append(stats)
        except Exception as e:
            results.append({'filename': filename, 'error': str(e)})
    return jsonify(results)

@app.route('/generate', methods=['POST'])
def generate():
    if 'file' in request.files:
        file = request.files['file']
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
    else:
        filename = request.form.get('filename')
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename) if filename else TRAINING_DATA_PATH

    year = request.form.get('year')
    month = request.form.get('month')

    if not os.path.exists(filepath):
        return render_template('index.html', error='No dataset available. Please upload an Excel file first.')

    month_name = ''
    if month and year:
        try:
            month_name = datetime(1900, int(month), 1).strftime('%B').upper() + f' {year}'
        except Exception:
            month_name = ''
    heading = f'CODIFICATION SUMMARY FOR THE MONTH OF {month_name}' if month_name else 'CODIFICATION SUMMARY'

    try:
        report = generate_report(filepath)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename_base = f'report_{ts}'
        excel_path = build_excel_from_report_data(report, heading, excel_filename_base)
        excel_filename = os.path.basename(excel_path)
        return render_template('report.html', data=report, now=datetime.now(),
                               heading=heading, last_report_filename=excel_filename)
    except Exception as ex:
        return render_template('index.html', error=f'Report generation failed: {str(ex)}')

@app.route('/generate_multi', methods=['POST'])
def generate_multi():
    """Generate reports for multiple files with per-file month/year settings."""
    data = request.json

    per_file = data.get('per_file', None)
    if per_file is None:
        filenames = data.get('filenames', [])
        year  = data.get('year', '')
        month = data.get('month', '')
        per_file = [{'filename': fn, 'year': year, 'month': month} for fn in filenames]

    results = []
    for entry in per_file:
        filename = entry.get('filename')
        year     = entry.get('year', '')
        month    = entry.get('month', '')

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename) if filename else TRAINING_DATA_PATH
        if not os.path.exists(filepath):
            results.append({'filename': filename, 'error': 'File not found'})
            continue

        month_name = ''
        if month and year:
            try:
                month_name = datetime(1900, int(month), 1).strftime('%B').upper() + f' {year}'
            except Exception:
                month_name = ''
        heading = f'CODIFICATION SUMMARY FOR THE MONTH OF {month_name}' if month_name else f'CODIFICATION SUMMARY — {filename}'

        try:
            report_data = generate_report(filepath)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_name = (filename or 'training').replace('.xlsx','').replace('.xls','').replace(' ','_')
            excel_filename_base = f'report_{safe_name}_{ts}'
            excel_path = build_excel_from_report_data(report_data, heading, excel_filename_base)
            excel_filename = os.path.basename(excel_path)

            rows_html = []
            totals = {'codified':0,'fwd':0,'nsn':0,'returned':0}
            serial = 1
            for dpsu, items in report_data.items():
                for idx, item in enumerate(items):
                    totals['codified'] += item['Total_Codified']
                    totals['fwd']      += item['Fwd_DCA']
                    totals['nsn']      += item['NSN']
                    totals['returned'] += item['Returned']
                    rows_html.append({
                        'serial': serial,
                        'dpsu': dpsu if idx == 0 else '',
                        'dpsu_rowspan': len(items) if idx == 0 else 0,
                        'equipment': item['Equipment'],
                        'total_codified': item['Total_Codified'],
                        'fwd_dca': item['Fwd_DCA'],
                        'nsn': item['NSN'],
                        'returned': item['Returned'],
                    })
                    serial += 1

            results.append({
                'filename': filename,
                'heading': heading,
                'period_month': month,
                'period_year': year,
                'excel_filename': excel_filename,
                'rows': rows_html,
                'totals': totals,
                'serial_end': serial,
            })
        except Exception as ex:
            results.append({'filename': filename, 'error': str(ex)})

    return jsonify(results)

@app.route('/download/<filename>')
def download(filename):
    path = os.path.join(app.config['REPORT_FOLDER'], filename)
    if not os.path.exists(path):
        return 'File not found', 404
    return send_file(os.path.abspath(path), as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/training_stats')
def training_stats():
    if not os.path.exists(TRAINING_DATA_PATH):
        return jsonify({'total': 0, 'forwarded': 0, 'nsn_allotted': 0,
                        'returned': 0, 'pending': 0,
                        'by_dpsu': {}, 'by_ncb': {}, 'by_equipment': {},
                        'avg_mrc': 0, 'avg_processing_days': 0,
                        'dpsu_list': [], 'total_rows': 0})
    df = pd.read_excel(TRAINING_DATA_PATH)
    engine = DataEngine(df)
    stats = engine.summary(df)
    dpsu_list = sorted(df['DPSU'].dropna().unique().tolist()) if 'DPSU' in df.columns else []
    stats['dpsu_list'] = dpsu_list
    stats['total_rows'] = len(df)
    return jsonify(stats)

if __name__ == '__main__':
    app.run(debug=True, port=5000)